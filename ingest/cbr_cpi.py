"""Официальная инфляция ИПЦ, % г/г (Росстат/ЦБ) → SQLite. Блок «Результат».

Источник: ЦБ, страница «Инфляция и ключевая ставка» cbr.ru/hd_base/infl/.
Под капотом — Excel-экспорт UniDbQuery:
  GET https://www.cbr.ru/Queries/UniDbQuery/DownloadExcel/132934
      ?Posted=True&FromDate=MM/DD/YYYY&ToDate=MM/DD/YYYY
  Лист «Инфляция и ключевая ставка»; колонки:
    Дата (MM.YYYY) | Ключевая ставка, % годовых | Инфляция, % г/г | Цель по инфляции
  Берём «Инфляция, % г/г». Частота — месячная (M).

⚠ ID запроса (132934) — это UniDbQuery table id, он может «уплыть». Если 404 —
открыть cbr.ru/hd_base/infl/ и найти актуальную ссылку DownloadExcel/<id>.
Формат/ID подтверждены напрямую 2026-06-23 (последнее: 05.2026 = 5,31% г/г).
К ЦБ — в обход прокси (cbr_session). openpyxl + сессия из ingest.
"""
from __future__ import annotations

import sys
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from db import DB_PATH, upsert_series, write_observations  # noqa: E402
from ingest import cbr_session  # noqa: E402

CPI_EXCEL_URL = "https://www.cbr.ru/Queries/UniDbQuery/DownloadExcel/132934"
SOURCE_URL = "https://www.cbr.ru/hd_base/infl/"
SHEET = "Инфляция и ключевая ставка Банк"
DEFAULT_FROM = "2024-01-01"


def _mmddyyyy(iso: str) -> str:
    return datetime.fromisoformat(iso).strftime("%m/%d/%Y")


def fetch_cpi_yoy(from_date: str = DEFAULT_FROM, to_date: str | None = None) -> list[tuple[str, float]]:
    """Месячный ряд ИПЦ % г/г → [(date 'YYYY-MM-01', value)] по возрастанию даты."""
    to_date = to_date or date.today().isoformat()
    params = {"Posted": "True", "FromDate": _mmddyyyy(from_date), "ToDate": _mmddyyyy(to_date)}
    resp = cbr_session().get(CPI_EXCEL_URL, params=params, timeout=60)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows: list[tuple[str, float]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):  # строка 1 — заголовок
        d, infl = r[0], r[2]  # колонка 2 (0-based) = «Инфляция, % г/г»
        if not isinstance(d, str) or not isinstance(infl, (int, float)) or isinstance(infl, bool):
            continue
        mm, yyyy = d.split(".")  # 'MM.YYYY'
        rows.append((f"{yyyy}-{mm}-01", float(infl)))
    if not rows:
        raise RuntimeError("ЦБ вернул пустой ряд ИПЦ — проверить ID запроса / формат")
    rows.sort(key=lambda x: x[0])
    return rows


def ingest(from_date: str = DEFAULT_FROM, to_date: str | None = None, db_path: Path | str = DB_PATH) -> int:
    rows = fetch_cpi_yoy(from_date, to_date)
    upsert_series(
        "cpi_yoy",
        title="Официальная инфляция ИПЦ, % г/г",
        unit="% г/г",
        freq="M",
        panel="result",
        source_name="Росстат / Банк России",
        source_url=SOURCE_URL,
        notes="cbr.ru/hd_base/infl/ (DownloadExcel/132934). Цель ЦБ — 4%. "
              "Запрос в обход прокси (РФ-IP).",
        db_path=db_path,
    )
    return write_observations("cpi_yoy", rows, flag="final", db_path=db_path)


if __name__ == "__main__":
    n = ingest()
    print(f"cpi_yoy: записано {n} наблюдений в {DB_PATH}")
