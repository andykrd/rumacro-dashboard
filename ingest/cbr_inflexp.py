"""Наблюдаемая инфляция и инфляционные ожидания населения (ИнФОМ) → SQLite.

Источник: ЦБ, отчёты «Инфляционные ожидания и потребительские настроения».
Каждый месячный отчёт публикуется и в xlsx, и в нём есть лист «Данные для графиков»
с ПОЛНЫМ месячным рядом (исходные данные графиков), а не только 3 последними месяцами.

  Лендинг (список отчётов, новейший сверху):
    https://www.cbr.ru/analytics/dkp/inflationary_expectations/
  Берём ссылку на новейший xlsx (…/Collection/File/<id>/Infl_exp_YY-MM.xlsx),
  на листе «Данные для графиков» — строки с точными метками:
    «наблюдаемая инфляция»  → observed_infl  (медиана, % за предыдущие 12 мес)
    «ожидаемая инфляция»    → infl_expect    (медиана, % на 12 мес вперёд)
  Даты — в строке выше блока; значения берём там, где заголовок-дата И число.
  Частота — месячная (M). Ряд начинается ~с 2024-01 (свежие отчёты дают ~29 мес).

Формат подтверждён напрямую 2026-06-23 (хвост: наблюдаемая 15.08, ожидаемая 13.02 за 05.2026).
⚠ File id xlsx меняется каждый месяц — поэтому НЕ хардкодим, а берём с лендинга.
К ЦБ — в обход прокси (cbr_session). openpyxl + сессия из ingest.
"""
from __future__ import annotations

import re
import sys
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from db import DB_PATH, upsert_series, write_observations  # noqa: E402
from ingest import cbr_session  # noqa: E402

LANDING = "https://www.cbr.ru/analytics/dkp/inflationary_expectations/"
BASE = "https://www.cbr.ru"
SHEET = "Данные для графиков"
# series_id → точная метка строки на листе (нижний регистр, без подгрупп «среди тех, кто…»)
LABELS = {"observed_infl": "наблюдаемая инфляция", "infl_expect": "ожидаемая инфляция"}

_META = {
    "observed_infl": dict(
        title="Наблюдаемая инфляция (опрос ИнФОМ)",
        unit="% (медиана, за 12 мес)",
        notes="Лист «Данные для графиков», строка «наблюдаемая инфляция». "
              "Разрыв gap = observed − cpi_yoy (на лету).",
    ),
    "infl_expect": dict(
        title="Инфляционные ожидания населения на год (ИнФОМ)",
        unit="% (медиана, на 12 мес вперёд)",
        notes="Лист «Данные для графиков», строка «ожидаемая инфляция».",
    ),
}


def _norm(s: object) -> str:
    return " ".join(str(s).split()).lower() if s is not None else ""


def _latest_xlsx_url(session) -> str:
    """Ссылка на новейший xlsx-отчёт ИнФОМ (с лендинга, новейший — первый)."""
    html = session.get(LANDING, timeout=30).text
    found = re.findall(r"/Collection/Collection/File/\d+/Infl_exp_[\d-]+\.xlsx", html)
    if not found:
        raise RuntimeError("на лендинге ИнФОМ не найдено ни одного xlsx — формат изменился?")
    return BASE + found[0]


def _extract(rows: list, label: str) -> list[tuple[str, float]]:
    """Достать ряд по метке: даты — из ближайшей строки-даты выше, значения — из строки метки."""
    ri = next((i for i, r in enumerate(rows) if _norm(r[0]) == label), None)
    if ri is None:
        raise RuntimeError(f"строка {label!r} не найдена на листе {SHEET!r} (метки ЦБ могли измениться)")
    di = next((i for i in range(ri - 1, -1, -1)
               if sum(isinstance(c, datetime) for c in rows[i]) >= 10), None)
    if di is None:
        raise RuntimeError(f"для {label!r} не найдена строка-дата")
    pts = [
        (d.date().isoformat(), float(v))
        for d, v in zip(rows[di], rows[ri])
        if isinstance(d, datetime) and isinstance(v, (int, float)) and not isinstance(v, bool)
    ]
    if not pts:
        raise RuntimeError(f"ряд {label!r} найден, но без числовых значений")
    pts.sort(key=lambda x: x[0])
    return pts


def fetch_inflexp() -> dict[str, list[tuple[str, float]]]:
    s = cbr_session()
    content = s.get(_latest_xlsx_url(s), timeout=60).content
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    return {sid: _extract(rows, lbl) for sid, lbl in LABELS.items()}


def ingest(db_path: Path | str = DB_PATH) -> dict[str, int]:
    data = fetch_inflexp()
    written: dict[str, int] = {}
    for sid, pts in data.items():
        upsert_series(
            sid, freq="M", panel="result",
            source_name="Банк России / ИнФОМ", source_url=LANDING,
            **_META[sid], db_path=db_path,
        )
        written[sid] = write_observations(sid, pts, flag="final", db_path=db_path)
    return written


if __name__ == "__main__":
    res = ingest()
    for sid, n in res.items():
        print(f"{sid}: записано {n} наблюдений")
    print(f"→ {DB_PATH}")
